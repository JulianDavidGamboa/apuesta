from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from datetime import datetime
from zoneinfo import ZoneInfo
import os
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dinero-app-2024')

DATABASE_URL = os.environ.get('DATABASE_URL')  # Neon en producción
USE_PG = bool(DATABASE_URL)

if USE_PG:
    import psycopg2
    import psycopg2.extras
else:
    import sqlite3
    DB = os.environ.get('DB_PATH', 'dinero.db')


# ── Conexión ─────────────────────────────────────────────────────────────────

def get_db():
    if USE_PG:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        return conn
    else:
        conn = sqlite3.connect(DB)
        conn.row_factory = sqlite3.Row
        return conn


def ph(n=1):
    """Retorna placeholders: %s para PG, ? para SQLite."""
    if USE_PG:
        return ', '.join(['%s'] * n)
    return ', '.join(['?'] * n)


def p(val=None):
    """Placeholder único."""
    return '%s' if USE_PG else '?'


def init_db():
    conn = get_db()
    cur = conn.cursor() if USE_PG else conn
    if USE_PG:
        cur.execute('''
            CREATE TABLE IF NOT EXISTS rondas (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                total_inicial REAL NOT NULL DEFAULT 0,
                total_ganado REAL NOT NULL DEFAULT 0,
                fecha TEXT NOT NULL
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS participantes (
                id SERIAL PRIMARY KEY,
                ronda_id INTEGER NOT NULL,
                nombre TEXT NOT NULL,
                porcentaje REAL NOT NULL,
                FOREIGN KEY (ronda_id) REFERENCES rondas(id)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS personas (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL UNIQUE
            )
        ''')
        # Migrate existing names
        cur.execute('''
            INSERT INTO personas (nombre)
            SELECT DISTINCT LOWER(TRIM(nombre)) FROM participantes
            ON CONFLICT (nombre) DO NOTHING
        ''')
        conn.commit()
        cur.close()
    else:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS rondas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                total_inicial REAL NOT NULL DEFAULT 0,
                total_ganado REAL NOT NULL DEFAULT 0,
                fecha TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS participantes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ronda_id INTEGER NOT NULL,
                nombre TEXT NOT NULL,
                porcentaje REAL NOT NULL,
                FOREIGN KEY (ronda_id) REFERENCES rondas(id)
            );
            CREATE TABLE IF NOT EXISTS personas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            );
            INSERT OR IGNORE INTO personas (nombre)
            SELECT DISTINCT LOWER(TRIM(nombre)) FROM participantes;
        ''')
    conn.close()


init_db()


# ── Helpers ───────────────────────────────────────────────────────────────────

def query(sql, params=(), one=False):
    conn = get_db()
    if USE_PG:
        cur = conn.cursor()
        cur.execute(sql, params)
        result = cur.fetchone() if one else cur.fetchall()
        cur.close()
        conn.close()
        return result
    else:
        result = conn.execute(sql, params)
        data = result.fetchone() if one else result.fetchall()
        conn.close()
        return data


def execute(sql, params=()):
    """Ejecuta escritura y retorna lastrowid."""
    conn = get_db()
    if USE_PG:
        cur = conn.cursor()
        cur.execute(sql, params)
        lastrowid = cur.fetchone()['id'] if 'RETURNING' in sql else None
        conn.commit()
        cur.close()
        conn.close()
        return lastrowid
    else:
        with conn:
            cur = conn.execute(sql, params)
            conn.commit()
        lastrowid = cur.lastrowid
        conn.close()
        return lastrowid


def execute_many(statements):
    """Ejecuta múltiples escrituras en una transacción."""
    conn = get_db()
    if USE_PG:
        cur = conn.cursor()
        for sql, params in statements:
            cur.execute(sql, params)
        conn.commit()
        cur.close()
    else:
        with conn:
            for sql, params in statements:
                conn.execute(sql, params)
            conn.commit()
    conn.close()


def _sql(sqlite_sql):
    """Convierte placeholders ? a %s si usamos PG."""
    if USE_PG:
        return sqlite_sql.replace('?', '%s')
    return sqlite_sql


def calcular_tabla(ronda, participantes):
    tabla = []
    for p_ in participantes:
        capital = ronda['total_inicial'] * p_['porcentaje'] / 100
        total_pool = ronda['total_ganado'] * p_['porcentaje'] / 100
        ganancia_neta = total_pool - capital
        total_queda = capital + ganancia_neta  # = total_pool
        tabla.append({
            'nombre': p_['nombre'],
            'porcentaje': p_['porcentaje'],
            'capital_dado': capital,
            'dinero_ganado': ganancia_neta,
            'total_queda': total_queda,
            'ganancia': ganancia_neta,
        })
    return tabla


# ── Rutas ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    rondas = query('SELECT * FROM rondas ORDER BY id DESC')
    return render_template('index.html', rondas=rondas)


@app.route('/nueva', methods=['GET', 'POST'])
def nueva_ronda():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        total_inicial = float(request.form.get('total_inicial', 0))
        total_ganado = float(request.form.get('total_ganado', 0))
        nombres_p = request.form.getlist('nombre_p')
        porcentajes = request.form.getlist('porcentaje')
        fecha = datetime.now(ZoneInfo('America/Bogota')).strftime('%Y-%m-%d %H:%M')

        if not nombre or not nombres_p:
            personas_list = query('SELECT nombre FROM personas ORDER BY nombre ASC')
            return render_template('nueva.html', error='Completa todos los campos.',
                                   personas=[r['nombre'] for r in personas_list])

        if USE_PG:
            ronda_id = execute(
                'INSERT INTO rondas (nombre, total_inicial, total_ganado, fecha) VALUES (%s, %s, %s, %s) RETURNING id',
                (nombre, total_inicial, total_ganado, fecha)
            )
        else:
            ronda_id = execute(
                'INSERT INTO rondas (nombre, total_inicial, total_ganado, fecha) VALUES (?, ?, ?, ?)',
                (nombre, total_inicial, total_ganado, fecha)
            )

        stmts = []
        for n, pct in zip(nombres_p, porcentajes):
            if n.strip() and pct:
                if USE_PG:
                    stmts.append(('INSERT INTO participantes (ronda_id, nombre, porcentaje) VALUES (%s, %s, %s)',
                                  (ronda_id, n.strip(), float(pct))))
                else:
                    stmts.append(('INSERT INTO participantes (ronda_id, nombre, porcentaje) VALUES (?, ?, ?)',
                                  (ronda_id, n.strip(), float(pct))))
        if stmts:
            execute_many(stmts)

        return redirect(url_for('ver_ronda', id=ronda_id))

    personas_list = query('SELECT nombre FROM personas ORDER BY nombre ASC')
    return render_template('nueva.html', error=None, personas=[r['nombre'] for r in personas_list])


@app.route('/ronda/<int:id>')
def ver_ronda(id):
    ronda = query(_sql('SELECT * FROM rondas WHERE id = ?'), (id,), one=True)
    if not ronda:
        return redirect(url_for('index'))
    participantes = query(
        _sql('SELECT * FROM participantes WHERE ronda_id = ? ORDER BY porcentaje DESC'), (id,)
    )
    tabla = calcular_tabla(ronda, participantes)
    return render_template('ronda.html', ronda=ronda, tabla=tabla)


@app.route('/ronda/<int:id>/eliminar', methods=['POST'])
def eliminar_ronda(id):
    clave = request.form.get('clave', '')
    clave_correcta = os.environ.get('DELETE_PASSWORD', '')
    if not clave_correcta or clave != clave_correcta:
        ronda = query(_sql('SELECT * FROM rondas WHERE id = ?'), (id,), one=True)
        participantes = query(_sql('SELECT * FROM participantes WHERE ronda_id = ? ORDER BY porcentaje DESC'), (id,))
        tabla = calcular_tabla(ronda, participantes)
        return render_template('ronda.html', ronda=ronda, tabla=tabla, error_clave=True)
    execute_many([
        (_sql('DELETE FROM participantes WHERE ronda_id = ?'), (id,)),
        (_sql('DELETE FROM rondas WHERE id = ?'), (id,)),
    ])
    return redirect(url_for('index'))


@app.route('/resumen')
def resumen():
    datos = query('''
        SELECT
            LOWER(TRIM(p.nombre)) AS nombre,
            COUNT(DISTINCT p.ronda_id) AS rondas,
            SUM(r.total_inicial * p.porcentaje / 100) AS total_capital,
            SUM((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS total_ganado,
            SUM(r.total_ganado * p.porcentaje / 100) AS total_recibido
        FROM participantes p
        JOIN rondas r ON p.ronda_id = r.id
        GROUP BY LOWER(TRIM(p.nombre))
        ORDER BY total_ganado DESC
    ''')
    return render_template('resumen.html', datos=datos)


@app.route('/estadisticas')
def estadisticas():
    # Globales
    globales = query('''
        SELECT
            COUNT(*) AS total_rondas,
            SUM(total_inicial) AS capital_total,
            SUM(total_ganado - total_inicial) AS ganancia_total,
            SUM(total_ganado) AS valor_final_total
        FROM rondas
    ''', one=True)

    mejor_ronda = query('''
        SELECT nombre, fecha,
               (total_ganado - total_inicial) AS ganancia
        FROM rondas ORDER BY ganancia DESC LIMIT 1
    ''', one=True)

    peor_ronda = query('''
        SELECT nombre, fecha,
               (total_ganado - total_inicial) AS ganancia
        FROM rondas ORDER BY ganancia ASC LIMIT 1
    ''', one=True)

    # Por persona
    por_persona = query('''
        SELECT
            LOWER(TRIM(p.nombre)) AS nombre,
            COUNT(DISTINCT p.ronda_id) AS rondas,
            SUM(r.total_inicial * p.porcentaje / 100) AS capital,
            SUM((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS ganancia_neta,
            SUM(r.total_ganado * p.porcentaje / 100) AS valor_final,
            MAX((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS mejor_ronda,
            MIN((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS peor_ronda,
            AVG((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS promedio_ronda,
            SUM(CASE WHEN (r.total_ganado - r.total_inicial) > 0 THEN 1 ELSE 0 END) AS rondas_ganadas,
            SUM(CASE WHEN (r.total_ganado - r.total_inicial) < 0 THEN 1 ELSE 0 END) AS rondas_perdidas
        FROM participantes p
        JOIN rondas r ON p.ronda_id = r.id
        GROUP BY LOWER(TRIM(p.nombre))
        ORDER BY MAX(r.fecha) DESC
    ''')

    # Datos para gráficas: ganancia por ronda por persona
    detalle_chart = query('''
        SELECT
            LOWER(TRIM(p.nombre)) AS nombre,
            r.nombre AS ronda_nombre,
            r.fecha,
            (r.total_ganado - r.total_inicial) * p.porcentaje / 100 AS ganancia
        FROM participantes p
        JOIN rondas r ON p.ronda_id = r.id
        ORDER BY LOWER(TRIM(p.nombre)), r.id ASC
    ''')

    chart_data = {}
    for row in detalle_chart:
        n = row['nombre']
        if n not in chart_data:
            chart_data[n] = {'labels': [], 'ganancias': [], 'acumulado': []}
        chart_data[n]['labels'].append(row['ronda_nombre'])
        g = round(float(row['ganancia']), 2)
        chart_data[n]['ganancias'].append(g)
        prev = chart_data[n]['acumulado'][-1] if chart_data[n]['acumulado'] else 0
        chart_data[n]['acumulado'].append(round(prev + g, 2))

    return render_template('estadisticas.html',
                           globales=globales,
                           mejor_ronda=mejor_ronda,
                           peor_ronda=peor_ronda,
                           por_persona=por_persona,
                           chart_data=json.dumps(chart_data))


@app.route('/calendario')
def calendario():
    rondas = query('SELECT * FROM rondas ORDER BY fecha ASC')
    rondas_por_fecha = {}
    for r in rondas:
        fecha_dia = r['fecha'][:10]
        if fecha_dia not in rondas_por_fecha:
            rondas_por_fecha[fecha_dia] = []
        rondas_por_fecha[fecha_dia].append({
            'id': r['id'],
            'nombre': r['nombre'],
            'total_inicial': r['total_inicial'],
            'total_ganado': r['total_ganado'],
            'ganancia': r['total_ganado'] - r['total_inicial'],
        })
    return render_template('calendario.html', rondas_por_fecha=json.dumps(rondas_por_fecha))


@app.route('/personas')
def personas():
    lista = query('SELECT * FROM personas ORDER BY nombre ASC')
    return render_template('personas.html', personas=lista, error_id=None)


@app.route('/personas/agregar', methods=['POST'])
def agregar_persona():
    nombre = request.form.get('nombre', '').strip().lower()
    if nombre:
        try:
            if USE_PG:
                execute('INSERT INTO personas (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING', (nombre,))
            else:
                execute('INSERT OR IGNORE INTO personas (nombre) VALUES (?)', (nombre,))
        except Exception:
            pass
    return redirect(url_for('personas'))


@app.route('/personas/<int:id>/eliminar', methods=['POST'])
def eliminar_persona(id):
    clave = request.form.get('clave', '')
    clave_correcta = os.environ.get('DELETE_PASSWORD', '')
    if not clave_correcta or clave != clave_correcta:
        lista = query('SELECT * FROM personas ORDER BY nombre ASC')
        return render_template('personas.html', personas=lista, error_id=id)
    execute(_sql('DELETE FROM personas WHERE id = ?'), (id,))
    return redirect(url_for('personas'))


@app.route('/api/personas')
def api_personas():
    lista = query('SELECT nombre FROM personas ORDER BY nombre ASC')
    return jsonify([r['nombre'] for r in lista])


@app.route('/exportar')
def exportar_excel():
    rondas = query('SELECT * FROM rondas ORDER BY fecha DESC')
    participantes = query('SELECT * FROM participantes')

    part_por_ronda = {}
    for p_ in participantes:
        part_por_ronda.setdefault(p_['ronda_id'], []).append(p_)

    wb = openpyxl.Workbook()
    header_fill = PatternFill('solid', fgColor='212529')
    header_font = Font(bold=True, color='FFFFFF')

    # Hoja 1: Rondas
    ws1 = wb.active
    ws1.title = 'Rondas'
    for col, h in enumerate(['Ronda', 'Fecha', 'Total Inicial ($)', 'Total Ganado ($)', 'Ganancia ($)'], 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for r in rondas:
        ws1.append([r['nombre'], r['fecha'], r['total_inicial'], r['total_ganado'], r['total_ganado'] - r['total_inicial']])
    for col in ws1.iter_cols(min_row=2, min_col=3, max_col=5):
        for cell in col:
            cell.number_format = '"$"#,##0'
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 18
    for c in ['C', 'D', 'E']:
        ws1.column_dimensions[c].width = 18

    # Hoja 2: Detalle
    ws2 = wb.create_sheet('Detalle')
    for col, h in enumerate(['Ronda', 'Fecha', 'Persona', 'Porcentaje (%)', 'Capital dado ($)', 'Ganado ($)', 'Total ($)'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for r in rondas:
        for p_ in part_por_ronda.get(r['id'], []):
            capital = r['total_inicial'] * p_['porcentaje'] / 100
            ganado  = r['total_ganado']  * p_['porcentaje'] / 100
            ws2.append([r['nombre'], r['fecha'], p_['nombre'], p_['porcentaje'], capital, ganado, capital + ganado])
    for col in ws2.iter_cols(min_row=2, min_col=5, max_col=7):
        for cell in col:
            cell.number_format = '"$"#,##0'
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 16
    for c in ['E', 'F', 'G']:
        ws2.column_dimensions[c].width = 18

    # Hoja 3: Resumen
    ws3 = wb.create_sheet('Resumen por persona')
    for col, h in enumerate(['Persona', 'Rondas', 'Capital total ($)', 'Total ganado ($)', 'Total recibido ($)', 'Ganancia neta ($)', 'ROI (%)'], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    resumen_data = query('''
        SELECT LOWER(TRIM(p.nombre)) AS nombre, COUNT(DISTINCT p.ronda_id) AS rondas,
               SUM(r.total_inicial * p.porcentaje / 100) AS total_capital,
               SUM((r.total_ganado - r.total_inicial) * p.porcentaje / 100) AS total_ganado,
               SUM(r.total_ganado * p.porcentaje / 100) AS total_recibido
        FROM participantes p JOIN rondas r ON p.ronda_id = r.id
        GROUP BY LOWER(TRIM(p.nombre)) ORDER BY total_ganado DESC
    ''')
    for d in resumen_data:
        ganancia_neta = d['total_recibido'] - d['total_capital']
        roi = (ganancia_neta / d['total_capital'] * 100) if d['total_capital'] else 0
        ws3.append([d['nombre'], d['rondas'], d['total_capital'], d['total_ganado'], d['total_recibido'], ganancia_neta, roi])
    for col in ws3.iter_cols(min_row=2, min_col=3, max_col=6):
        for cell in col:
            cell.number_format = '"$"#,##0'
    for cell in ws3['G'][1:]:
        cell.number_format = '0.0"%"'
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 10
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws3.column_dimensions[c].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'kuazimides_{datetime.now(ZoneInfo('America/Bogota')).strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True)
